import email
import imaplib
import json
import glob
from openpyxl import load_workbook
global workbook
from selenium.webdriver.common.by import By
import pytesseract
from PIL import Image
from PyPDF2 import PdfReader
import os
from robot.api.deco import keyword
from selenium import webdriver

def fetch_user_details_by_number(file_path, target_id):
    global workbook
    try:
        workbook = load_workbook(file_path)
        sheet = workbook.active
        header = [cell.value for cell in sheet[1]]

        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0] == target_id:
                data_dict = {}
                for col_name, value in zip(header, row):
                    if ',' in str(value):
                        data_dict[col_name] = [item.strip() for item in value.split(',')]
                    else:
                        data_dict[col_name] = value
                return data_dict
    except Exception as e:
        print(f"Error reading Excel file: {e}")
    finally:
        try:
            workbook.close()
        except NameError:
            pass
    return None


def fetch_data_by_id(file_path, target_id):
    global workbook
    try:
        workbook = load_workbook(file_path)
        sheet = workbook.active
        header = [cell.value for cell in sheet[1]]

        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0] == target_id:
                data_dict = {}
                for col_name, value in zip(header, row):
                    if ',' in str(value):
                        data_dict[col_name] = [item.strip() for item in value.split(',')]
                    else:
                        data_dict[col_name] = value
                return data_dict
    except Exception as e:
        print(f"Error reading Excel file: {e}")
    finally:
        try:
            workbook.close()
        except NameError:
            pass
    return None


def write_to_json_file(data, file_path):
    with open(file_path, 'w') as f:
        json.dump(data, f, indent=4)


def read_from_json_file(file_path):
    with open(file_path, 'r') as f:
        return json.load(f)


def check_download_and_return_filepath(timestamp):
    # user_name = getpass.getuser()
    script_dir = os.path.dirname(__file__)
    default_download_path = os.path.join(script_dir, '..', '..', 'Downloads', 'Files', '*')
    # default_download_path = os.path.abspath(r'..\\Download\Files\*')
    files_list = glob.glob(default_download_path)
    print(files_list)
    for i in files_list:
        if timestamp < os.path.getctime(i):
            print(timestamp)
            os.path.getctime(i)
            return i
    return False


def get_environment_data(filename):
    try:
        with open(filename, 'r') as f:
            data = json.load(f)
    except (FileNotFoundError, json.JSONDecodeError) as e:
        raise Exception(f"Error processing JSON file: {filename}") from e
    return data


def is_valid_email(email):
    # Regular expression pattern for validating email addresses
    pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'

    # Check if the email matches the pattern
    if re.match(pattern, email):
        return True
    else:
        return False

def fetch_user_data_by_id(file_path, target_id):
    global workbook
    try:
        workbook = load_workbook(file_path)
        sheet = workbook.active
        header = [cell.value for cell in sheet[1]]

        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0] == target_id:
                data_dict = {}
                for col_name, value in zip(header, row):
                    if ',' in str(value):
                        data_dict[col_name] = [item.strip() for item in value.split(',')]
                    else:
                        data_dict[col_name] = value
                return data_dict
    except Exception as e:
        print(f"Error reading Excel file: {e}")
    finally:
        try:
            workbook.close()
        except NameError:
            pass
    return None

def delete_all_emails_inbox(server, port, email_address, password):
    mail = imaplib.IMAP4_SSL(server, port)
    mail.login(email_address, password)
    mail.select('inbox')

    # Search for all emails in the inbox
    result, data = mail.search(None, "ALL")
    email_ids = data[0].split()

    # Delete all emails
    for email_id in email_ids:
        mail.store(email_id, '+FLAGS', '\\Deleted')
    mail.expunge()
    mail.logout()
    print("All emails deleted.")

def search_and_fetch_email(server, port, email_address, password, subject):
    # Connect to the email server
    mail = imaplib.IMAP4_SSL(server, port)
    mail.login(email_address, password)
    mail.select('inbox')

    # Search for emails with the specified subject
    result, data = mail.search(None, f'(SUBJECT "{subject}")')
    email_ids = data[0].split()

    # Fetch the latest email and extract the OTP from its body
    if email_ids:
        latest_email_id = email_ids[-1]
        result, data = mail.fetch(latest_email_id, '(RFC822)')
        raw_email = data[0][1].decode('utf-8', errors='replace')  # Decode bytes to string

        email_message = email.message_from_string(raw_email)  # Use message_from_string to parse email content

        # Extract the OTP from the email body
        email_body = ""
        received_time = email_message.get("Date")
        sender = email_message.get("From")
        if email_message.is_multipart():
            for part in email_message.walk():
                if part.get_content_type() == "text/plain":
                    email_body += part.get_payload(decode=True).decode('utf-8', errors='replace')
        else:
            email_body = email_message.get_payload(decode=True).decode('utf-8', errors='replace')

        return email_body, received_time, sender
    return None, None, None


def check_url(link):
    try:
        response = requests.get(link, timeout=5)  # Set timeout to 5 seconds
        response_text = response.text  # Get the response content as text
        if response.status_code >= 400:
            return f"{link} ---> {response.reason}\nResponse Content:\n{response_text}"
        else:
            return f"{link} ------> {response.reason}\nResponse Content:\n{response_text}"
    except requests.exceptions.RequestException as e:
        return f"{link} ---> Request failed: {e}"

def get_all_links(url):
        driver = webdriver.Chrome()  # You can configure this to use the appropriate WebDriver
        driver.get(url)
        links = driver.find_elements(By.TAG_NAME, 'a')
        hrefs = [link.get_attribute('href') for link in links if link.get_attribute('href')]
        driver.quit()
        return hrefs

def resize_image1(image_path, width, height, output_path):

    image = Image.open(image_path)
    resized_image = image.resize((int(width), int(height)))
    resized_image.save(output_path)
    return output_path

def crop_image(image_path1, output_path1, target_width, target_height):

    target_width = int(target_width)
    target_height = int(target_height)

    with Image.open(image_path1) as img:
        width, height = img.size

        # Calculate cropping coordinates
        left = int((width - target_width) / 2)
        top = int((height - target_height) / 2)
        right = int((width + target_width) / 2)
        bottom = int((height + target_height)/2)

        # Perform the crop operation
        cropped_img = img.crop((left, top, right, bottom))
        cropped_img.save(output_path1)

    return output_path1


def crop_image_header_footer(image_path, output_path, header_height, footer_height):
    # Open the image file
    with Image.open(image_path) as img:
        width, height = img.size
        header_height = int(header_height)
        footer_height = int(footer_height)
        # Define the crop box (left, upper, right, lower)
        # Assuming header and footer heights are known
        crop_box = (0, header_height, width, height - footer_height)

        # Crop the image
        cropped_img = img.crop(crop_box)

        # Save the cropped image
        cropped_img.save(output_path)
    return output_path



def crop_image1(image_path, output_path, target_width, target_height, top_crop=50, bottom_crop=50):
    target_width = int(target_width)
    target_height = int(target_height)

    with Image.open(image_path) as img:
        width, height = img.size

        # Calculate the cropping coordinates
        left = int((width - target_width) / 2)
        top = top_crop
        right = int((width + target_width) / 2)
        bottom = bottom_crop

        # Perform the crop operation
        cropped_img = img.crop((left, top, right, bottom))
        cropped_img.save(output_path)

    return output_path
@keyword
def extract_response_values(response, key):
    values = []
    def extract(obj, key):
        try:
            if isinstance(obj, dict):
                for k, v in obj.items():
                    if k == key:
                        values.append(v)
                    elif isinstance(v, (dict, list)):
                        extract(v, key)
            elif isinstance(obj, list):
                for item in obj:
                    extract(item, key)
        except Exception as e:
            print(f"An error occurred: {e}")
    try:
        extract(response, key)
    except Exception as e:
        print(f"An error occurred while processing the response: {e}")
    return values

def replace_am_pm(datetime_str):
    return datetime_str.replace('AM', 'am').replace('PM', 'pm')


def replace_am_pm(datetime_str):
    return datetime_str.replace('AM', 'am').replace('PM', 'pm')
@keyword
def remove_invalid_escapes(input_string):
    try:
        invalid_escape_pattern = re.compile(r'\\(?!["\\/bfnrtu])')
        corrected_string = invalid_escape_pattern.sub('', input_string)
        return corrected_string

    except Exception as e:
        print(f"An error occurred: {e}")
        return None

@keyword
def get_latest_image(directory_path):
    directory_path = os.path.abspath(directory_path)
    if not os.path.exists(directory_path):
        raise FileNotFoundError(f"Directory does not exist: {directory_path}")

    files = [os.path.join(directory_path, f) for f in os.listdir(directory_path) if f.endswith('.png')]
    if not files:
        raise FileNotFoundError("No Image found in the specified directory.")

    latest_file = max(files, key=os.path.getmtime)
    return latest_file

@keyword
def get_text_from_image(image_path):
        """Converts an image to text using Tesseract OCR."""
        image = Image.open(image_path)
        text = pytesseract.image_to_string(image)
        print(text)
        return text

def extract_pdf_details(pdf_path):
    reader = PdfReader(pdf_path)
    text = ''

    for page in reader.pages:
        text += page.extract_text()

    return {
        'text': text
    }

def extract_field(text, field_name):
    for line in text.split('\n'):
        if field_name in line:
            return line.split(':', 1)[1].strip()
    raise ValueError(f'{field_name} not found in the text')

def clean_text(text):
    return ''.join(char for char in text if char.isprintable())

def fetch_user_data_by_number(file_path, target_id):
    try:
        workbook = load_workbook(file_path)
        sheet = workbook.active
        header = [cell.value for cell in sheet[1]]

        for row in sheet.iter_rows(min_row=2, values_only=True):
            if str(row[4]) == str(target_id):  # Assuming mobile_no is the 5th column
                data_dict = {}
                for col_name, value in zip(header, row):
                    if ',' in str(value):
                        data_dict[col_name] = [item.strip() for item in value.split(',')]
                    else:
                        data_dict[col_name] = value
                return data_dict

    except Exception as e:
        print(f"Error reading Excel file: {e}")
    finally:
        try:
            workbook.close()
        except Exception as e:
            print(f"Error closing the workbook: {e}")

    return None